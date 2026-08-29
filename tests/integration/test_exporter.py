from openpyxl import load_workbook
from pdf_to_excel.excel.exporter import export_tables
from pdf_to_excel.models import ExtractedTable, SourceType, TableCell


def test_exports_real_workbook(tmp_path) -> None:
    table = ExtractedTable(
        1,
        1,
        [
            [TableCell(0, 0, "БМ"), TableCell(0, 1, "Име")],
            [TableCell(1, 0, "12"), TableCell(1, 1, "Марко")],
        ],
        SourceType.DIGITAL,
    )
    output = tmp_path / "result.xlsx"
    export_tables([table], output)
    sheet = load_workbook(output).active
    assert sheet["A2"].value == 12
    assert sheet["B2"].value == "Марко"


def test_structured_workbook_uses_serbian_labels(tmp_path) -> None:
    from pdf_to_excel.models import EquipmentItem, OutputMode, ReversDocument

    document = ReversDocument(tmp_path / "ulaz.pdf", 1)
    document.person_name = "Marko Marković"
    document.equipment_items.append(EquipmentItem(equipment_type="Računar"))

    output = export_tables([], tmp_path / "strukturirano.xlsx", [document], OutputMode.STRUCTURED)
    workbook = load_workbook(output)

    assert workbook.sheetnames == ["Oprema", "Dokument"]
    assert [cell.value for cell in workbook["Oprema"][1]] == [
        "Izvorna datoteka",
        "Stranica",
        "Osoba",
        "JMBG",
        "Organizaciona jedinica",
        "Redni broj",
        "Vrsta opreme",
        "Model",
        "Količina",
        "Serijski broj",
        "Inventarski broj",
        "Datum primopredaje",
        "Opremu predao",
        "Opremu primio",
        "Pouzdanost",
    ]
