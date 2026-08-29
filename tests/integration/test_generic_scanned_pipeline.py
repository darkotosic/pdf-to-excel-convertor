from pathlib import Path

import fitz
import pytest

from openpyxl import load_workbook

from pdf_to_excel.conversion.pipeline import ConversionPipeline
from pdf_to_excel.pdf.renderer import render_page_safe
from fixture_factory import create_clean_scanned_ruled_pdf

from pdf_to_excel.models import (
    BoundingBox,
    ConversionOptions,
    DocumentWord,
    OCRMode,
    OutputMode,
    WordSource,
)


def test_generic_scanned_ruled_table_flows_through_real_pipeline(tmp_path, monkeypatch) -> None:
    # OCR is made deterministic because CI need not have a system Tesseract installation;
    # rendering, OpenCV grid detection, cell assignment, pipeline, and XLSX export are real.
    rows = [
        ["Name", "City", "Phone"],
        ["Marko Markovic", "Ruma", "0641234567"],
        ["Jovan Jovanovic", "Sabac", "0631234567"],
        ["Djordje Petrovic", "Cacak", "0621234567"],
    ]
    xs = [120, 520, 820]
    ys = [190, 315, 440, 565]
    words = [
        DocumentWord(
            text,
            BoundingBox(xs[column], ys[row], xs[column] + 220, ys[row] + 35),
            0.99,
            1,
            WordSource.OCR,
        )
        for row, values in enumerate(rows)
        for column, text in enumerate(values)
    ]
    monkeypatch.setattr(
        "pdf_to_excel.conversion.page_processor.PageProcessor._ensure_ocr", lambda self: object()
    )
    monkeypatch.setattr(
        "pdf_to_excel.conversion.page_processor.PageProcessor._extract_ocr",
        lambda self, image, page_number: words,
    )
    monkeypatch.setattr(
        "pdf_to_excel.conversion.page_processor.correct_orientation", lambda image: image
    )
    monkeypatch.setattr("pdf_to_excel.conversion.page_processor.deskew", lambda image: image)

    fixture = create_clean_scanned_ruled_pdf(tmp_path / "generic_scanned_ruled_clean.pdf")
    output = tmp_path / "generic-scanned.xlsx"
    result = ConversionPipeline().convert(
        ConversionOptions(
            fixture,
            output,
            ocr_mode=OCRMode.ALWAYS,
            languages=("eng",),
            dpi=100,
            output_mode=OutputMode.BOTH,
        )
    )

    assert len(result.tables) == 1
    assert len(result.tables[0].rows) == 4
    assert len(result.tables[0].rows[0]) == 3
    assert [cell.text for cell in result.tables[0].rows[0]] == ["Name", "City", "Phone"]
    sheet = load_workbook(result.output_path)["Stranica 1 Tabela 1"]
    assert sheet.max_row == 4
    assert sheet.max_column == 3
    assert sheet["C2"].value == "0641234567"


def test_malformed_icc_scan_is_not_accepted_as_blank_content(tmp_path) -> None:
    from pdf_to_excel.exceptions import PdfRenderError

    fixture = Path(__file__).parents[1] / "fixtures" / "malformed_icc_scan.pdf"
    with fitz.open(fixture) as document, pytest.raises(PdfRenderError, match="suspicious"):
        render_page_safe(document[0], dpi=100)
