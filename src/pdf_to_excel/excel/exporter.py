from pathlib import Path
import os
import tempfile
from openpyxl import Workbook, load_workbook
from pdf_to_excel.models import ExtractedTable, OutputMode, ReversDocument
from pdf_to_excel.text.type_inference import infer_value
from .formatting import format_sheet
from .security import safe_excel_text


def export_tables(
    tables: list[ExtractedTable],
    destination: Path,
    documents: list[ReversDocument] | None = None,
    output_mode: OutputMode = OutputMode.PRESERVE_TABLES,
    include_empty_rows: bool = False,
) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        destination = available_output_path(destination)
    descriptor, name = tempfile.mkstemp(
        prefix=f".{destination.stem}-", suffix=".xlsx", dir=destination.parent
    )
    os.close(descriptor)
    temporary = Path(name)
    workbook = Workbook()
    active = workbook.active
    if active is not None:
        workbook.remove(active)
    documents = documents or []
    if not tables and not documents:
        workbook.create_sheet("Nema tabela").append(["Nije pronađen tabelarni sadržaj."])
    if documents and output_mode in (OutputMode.STRUCTURED, OutputMode.BOTH):
        _add_structured_sheets(workbook, documents, include_empty_rows)
    for table in tables if output_mode in (OutputMode.PRESERVE_TABLES, OutputMode.BOTH) else []:
        sheet = workbook.create_sheet(
            _unique_title(
                workbook.sheetnames,
                f"Stranica {table.page_number} Tabela {table.table_index}",
            )
        )
        for row in table.rows:
            values = []
            for cell in row:
                inferred = infer_value(cell.text)
                values.append(
                    inferred if not isinstance(inferred, str) else safe_excel_text(inferred)
                )
            sheet.append(values)
        format_sheet(sheet)
    try:
        workbook.save(temporary)
        verified = load_workbook(temporary, read_only=True)
        try:
            if not verified.sheetnames:
                raise ValueError("Napravljena radna sveska ne sadrži nijedan list")
        finally:
            verified.close()
        os.replace(temporary, destination)
    finally:
        temporary.unlink(missing_ok=True)
    return destination


def _add_structured_sheets(
    workbook: Workbook, documents: list[ReversDocument], include_empty_rows: bool
) -> None:
    equipment = workbook.create_sheet("Oprema")
    equipment.append(
        [
            "Izvorna datoteka",
            "Stranica",
            "Osoba",
            "JMBG",
            "Organizaciona jedinica",
            "Redni broj",
            "Vrsta opreme",
            "Model",
            "Količina",
            "Serijski broj",
            "Inventarski broj",
            "Datum primopredaje",
            "Opremu predao",
            "Opremu primio",
            "Pouzdanost",
        ]
    )
    for document in documents:
        items = document.equipment_items if include_empty_rows else document.populated_items()
        for item in items:
            equipment.append(
                [
                    document.source_file.name,
                    document.page_number,
                    document.person_name,
                    document.person_identifier,
                    document.organization_unit,
                    item.item_number,
                    item.equipment_type,
                    item.model,
                    item.quantity,
                    item.serial_number,
                    item.inventory_number,
                    document.handover_date,
                    document.handed_over_by,
                    document.received_by,
                    item.confidence,
                ]
            )
    for row in range(2, equipment.max_row + 1):
        for column in (4, 10, 11):
            equipment.cell(row, column).number_format = "@"
    format_sheet(equipment)
    human = workbook.create_sheet("Dokument")
    for document in documents:
        human.append(["REVERS"])
        human.append(["Osoba", document.person_name])
        human.append(["JMBG", document.person_identifier])
        human.append(["Organizaciona jedinica", document.organization_unit])
        human.append(["Datum primopredaje", document.handover_date])
        human.append(["Opremu predao", document.handed_over_by])
        human.append(["Opremu primio", document.received_by])
        human.append([])
        human.append(
            [
                "Red. broj",
                "Vrsta računarske opreme",
                "Model",
                "Kol",
                "Serijski broj",
                "Inventarski broj",
            ]
        )
        for item in document.equipment_items if include_empty_rows else document.populated_items():
            human.append(
                [
                    item.item_number,
                    item.equipment_type,
                    item.model,
                    item.quantity,
                    item.serial_number,
                    item.inventory_number,
                ]
            )
    format_sheet(human)
    warnings = [(document, warning) for document in documents for warning in document.warnings]
    if warnings:
        review = workbook.create_sheet("Pregled")
        review.append(
            [
                "Ozbiljnost",
                "Šifra upozorenja",
                "Stranica",
                "Red",
                "Polje",
                "Izdvojena vrednost",
                "Pouzdanost",
                "Izvor",
                "Upozorenje",
            ]
        )
        for document, warning in warnings:
            review.append(
                [
                    warning.severity.value,
                    warning.code,
                    warning.page_number,
                    warning.row,
                    warning.field,
                    warning.value,
                    warning.confidence,
                    warning.source.value if warning.source else "",
                    warning.message,
                ]
            )
        format_sheet(review)


def available_output_path(destination: Path) -> Path:
    if not destination.exists():
        return destination
    index = 1
    while True:
        candidate = destination.with_name(f"{destination.stem} ({index}){destination.suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def _unique_title(existing: list[str], preferred: str) -> str:
    base = preferred[:31]
    title, index = base, 2
    while title in existing:
        suffix = f" ({index})"
        title, index = base[: 31 - len(suffix)] + suffix, index + 1
    return title
